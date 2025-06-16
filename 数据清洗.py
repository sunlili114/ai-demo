import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import json

def clean_value(value):
    """清理单个值中的空值和斜杠"""
    if value is None or value == '' or value == '/':
        return 'N/A'
    return str(value).strip()

def process_merged_cells(ws):
    """合并单元格映射生成器"""
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        main_value = clean_value(ws.cell(row=min_row, column=min_col).value)
        
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell_coord = f"{get_column_letter(col)}{row}"
                merged_map[cell_coord] = main_value
    return merged_map

def convert_excel_ultimate(input_path):
    try:
        wb = load_workbook(input_path)
        writer = pd.ExcelWriter('Final_Output.xlsx', engine='openpyxl')
        
        # ========== 通用处理函数 ==========
        def process_sheet(ws, config):
            """通用表单处理器"""
            merged_map = process_merged_cells(ws)
            sheet_data = []
            
            # [修改] 首先获取所有可能的参数名作为列
            param_names = []
            for row in range(config['param_start'], config['param_end'] + 1):
                param_col = 2 if config['sheet_name'] == 'E-Motorcycle' else 1
                param_name = ws.cell(row=row, column=param_col).value
                if param_name:
                    param_names.append(param_name)
            
            for col in range(config['start_col'], ws.max_column + 1):
                col_letter = get_column_letter(col)
                
                # 获取产品名称
                product_cell = f"{col_letter}{config['title_row']}"
                product_name = clean_value(merged_map.get(product_cell, ws[product_cell].value))
                if product_name == 'N/A':
                    continue

                # 替换产品名称中的空格
                product_name = product_name.replace(' ', '_')

                # [修改] 创建行数据，包含所有参数列
                row_data = {'名称': product_name}
                
                # [修改] 填充每个参数的值
                for row, param_name in enumerate(param_names, start=config['param_start']):
                    param_value = clean_value(merged_map.get(
                        f"{col_letter}{row}",
                        ws.cell(row=row, column=col).value
                    ))
                    row_data[param_name] = param_value
                
                # 提取产品说明
                desc = 'N/A'
                if 'desc_row' in config:
                    desc_cell = f"{col_letter}{config['desc_row']}"
                    desc = clean_value(merged_map.get(desc_cell, ws.cell(row=config['desc_row'], column=col).value))
                
                row_data['卖点'] = desc
                sheet_data.append(row_data)
            
            return sheet_data

        # ========== E-Motorcycle处理 ==========
        motor_config = {
            'sheet_name': 'E-Motorcycle',
            'title_row': 2,      # 产品名称在第2行
            'start_col': 3,      # 从C列开始
            'param_start': 4,    # 参数起始行
            'param_end': 22,     # 参数结束行
            'desc_row': 23       # 说明在第23行
        }
        motor_data = process_sheet(wb['E-Motorcycle'], motor_config)
        pd.DataFrame(motor_data).to_excel(writer, motor_config['sheet_name'], index=False)

        # ========== E-Scooter处理 ==========
        scooter_config = {
            'sheet_name': 'E-Scooter',
            'title_row': 1,      # 产品名称在第1行
            'start_col': 3,      # 从C列开始
            'param_start': 2,    # 参数起始行
            'param_end': 12     # 参数结束行
        }
        scooter_data = process_sheet(wb['E-Scooter'], scooter_config)
        pd.DataFrame(scooter_data).to_excel(writer, scooter_config['sheet_name'], index=False)

        # ========== E-Bike处理 ==========
        bike_config = {
            'sheet_name': 'E-Bike',
            'title_row': 1,      # 产品名称在第1行
            'start_col': 3,      # 从C列开始
            'param_start': 2,    # 参数起始行
            'param_end': 12      # 参数结束行
        }
        bike_data = process_sheet(wb['E-Bike'], bike_config)
        pd.DataFrame(bike_data).to_excel(writer, bike_config['sheet_name'], index=False)

        # ========== E-Tricycle特殊处理 ==========
        ws_tri = wb['E-Tricycle']
        tri_merged = process_merged_cells(ws_tri)
        tri_data = []
        
        # 定义两个产品区块
        sections = [
            {'start_row': 1, 'end_row': 1, 'param_end': 5, 'type': 'E-Leisure Tricycle'},
            {'start_row': 8, 'end_row': 8, 'param_end': 14, 'type': 'E-Cargo Tricycle'}
        ]
        
        for section in sections:
            for col in range(3, ws_tri.max_column + 1):
                col_letter = get_column_letter(col)
                
                # 获取产品名称
                product_cell = f"{col_letter}{section['start_row']}"
                product_name = tri_merged.get(product_cell, ws_tri[product_cell].value)
                if not product_name:
                    continue
                
                # 替换产品名称中的空格
                product_name = str(product_name).replace(' ', '_')     

                # 提取参数
                params = []
                for row in range(section['start_row'] + 1, section['param_end'] + 1):
                    param_name = ws_tri.cell(row=row, column=1).value
                    param_value = tri_merged.get(
                        f"{col_letter}{row}",
                        ws_tri.cell(row=row, column=col).value
                    )
                    param_value = clean_value(param_value)
                    if param_name:
                        params.append(f"{param_name}:{param_value}")
                
                tri_data.append({
                    '名称': product_name,
                    '参数': '\n'.join(params) or 'N/A',
                    '卖点': 'N/A'  # 使用N/A替代空卖点
                })
        
        pd.DataFrame(tri_data).to_excel(writer, 'E-Tricycle', index=False)

        writer.close()
        
        # 结果验证
        if os.path.exists('Final_Output.xlsx') and os.path.getsize('Final_Output.xlsx') > 1024:
            print("生成成功！文件大小：", os.path.getsize('Final_Output.xlsx'), "字节")
            return True
        return False

    except Exception as e:
        print(f"生成失败: {str(e)}")
        return False

# 执行转换
if convert_excel_ultimate('AI机器人产品素材提供(2).xlsx'):
    print("请检查 Final_Output.xlsx")
else:
    print("转换失败，请检查输入文件")